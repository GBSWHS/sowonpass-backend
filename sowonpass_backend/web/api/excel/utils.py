from io import BytesIO
from typing import Any, Set

import pandas as pd
from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from sowonpass_backend.db.models.user import UserModel


def parse_single_sheet(
    worksheet: Worksheet,
    required_columns: Set[str],
) -> list[dict[str, Any]]:
    headers = []
    for header_cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
        headers.append(header_cell.value)

    if not required_columns.issubset(set(headers)):
        missing = required_columns - set(headers)
        raise ValueError(f"Missing columns: {missing}")

    return [
        create_row_data(row, headers)
        for row in worksheet.iter_rows(min_row=2, values_only=False)
        if not all(cell.value is None for cell in row)
    ]


def create_row_data(row: Any, headers: list[Any]) -> dict[str, Any]:
    row_data = {}
    for idx, cell in enumerate(row):
        if cell.value is not None:
            row_data[headers[idx]] = cell.value
    row_data["colored"] = any(
        cel.fill.start_color.index not in {"00000000", "FFFFFFFF"} for cel in row
    )
    return row_data


def parse_excel_file(file: UploadFile, required_columns: Set[str]) -> dict[str, Any]:
    try:
        workbook = load_workbook(filename=file.file, data_only=True)
    except Exception as err:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(err))

    all_data = {}

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        try:
            all_data[sheet] = parse_single_sheet(worksheet, required_columns)
        except ValueError as value_err:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Error in sheet '{sheet}': {value_err}",
            )

    return all_data


def process_users(
    users: list[UserModel],
    required: set[str],
) -> list[dict[str, Any]]:
    return [{field: getattr(user, field, None) for field in required} for user in users]


def create_excel_file(data: list[dict[str, Any]], sheet_name: str = "users") -> BytesIO:
    df = pd.DataFrame(data)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, na_rep="NaN")

        for col_idx, column_name in enumerate(df.columns):
            writer.sheets[sheet_name].set_column(
                col_idx,
                col_idx,
                max(
                    df[column_name].astype(str).apply(len).max(),
                    len(column_name),
                )
                + 3,
            )

    output.seek(0)
    return output


def create_excel_file_from_user_data(
    users: list[UserModel],
    required_fields: set[str],
    default_rows: list[dict[str, Any]],
) -> BytesIO:
    processed_users = process_users(users, required_fields)
    data = processed_users if processed_users else default_rows
    return create_excel_file(data)
